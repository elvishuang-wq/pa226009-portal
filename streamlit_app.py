"""
PA226009 客戶入口網站
讀取 Google Sheet「PA226009_Customer_Data」裡的 BOM_Matrix / Purchase_Detail / Receipt_Detail，
用網址帶 Token 驗證客戶身分，不需要客戶登入 Google。

需求套件見 requirements.txt

網址格式：
    https://your-app.streamlit.app/?token=xxxxxxxx
"""

import base64
import pickle
import io

import streamlit as st
import pandas as pd
import gspread
from google.auth.transport.requests import Request
from st_aggrid import AgGrid, GridUpdateMode
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

st.set_page_config(page_title="PA226009 專案進度", layout="wide")

GOOGLE_SHEET_KEY = "1_kdF00Cf9xwwByjlubAdwhyTWcvVIHFspjElAGaB2P8"
SCOPES = ["https://www.googleapis.com/auth/spreadsheets.readonly"]


# ---------- Google 憑證：優先讀 Streamlit secrets（雲端部署用），本地開發時退回讀 token.pickle ----------
import os


@st.cache_resource
def get_credentials():
    creds = None
    secrets_path = os.path.join(os.getcwd(), ".streamlit", "secrets.toml")

    if os.path.exists(secrets_path):
        try:
            if "token_pickle_b64" in st.secrets:
                raw = base64.b64decode(st.secrets["token_pickle_b64"])
                creds = pickle.load(io.BytesIO(raw))
        except Exception:
            creds = None  # secrets.toml 存在但沒有這個 key，或格式有問題，退回讀本地檔案

    if creds is None:
        with open("token.pickle", "rb") as f:
            creds = pickle.load(f)

    if not creds.valid:
        if creds.expired and creds.refresh_token:
            creds.refresh(Request())
        else:
            st.error("Google 憑證已失效，需要重新產生 token.pickle 並更新部署設定。")
            st.stop()

    return creds


@st.cache_data(ttl=300)
def load_sheet_data():
    creds = get_credentials()
    gc = gspread.authorize(creds)
    sh = gc.open_by_key(GOOGLE_SHEET_KEY)

    matrix = pd.DataFrame(sh.worksheet("BOM_Matrix").get_all_records())
    purchase = pd.DataFrame(sh.worksheet("Purchase_Detail").get_all_records())
    receipt = pd.DataFrame(sh.worksheet("Receipt_Detail").get_all_records())
    tokens = pd.DataFrame(sh.worksheet("Customer_Tokens").get_all_records())
    products_info = pd.DataFrame(sh.worksheet("Products_Info").get_all_records())

    return matrix, purchase, receipt, tokens, products_info


def check_token(tokens_df, token):
    if token is None or token.strip() == "":
        return False
    tokens_df = tokens_df.copy()
    tokens_df["Token"] = tokens_df["Token"].astype(str).str.strip()
    match = tokens_df[tokens_df["Token"] == token.strip()]
    if match.empty:
        return False
    if "啟用" in match.columns:
        return str(match.iloc[0]["啟用"]).strip().upper() in ("Y", "TRUE", "1")
    return True


def build_excel(matrix_view, purchase_df, receipt_df, products_info_df, product_cols, product_info_map, summary_cols):
    wb = Workbook()
    ws = wb.active
    ws.title = "零件總表"

    header_font = Font(bold=True)
    center = Alignment(horizontal="center", vertical="center", wrap_text=True)

    ws.merge_cells("A1:A5")
    ws["A1"] = "Description"
    ws.merge_cells("B1:B5")
    ws["B1"] = "Specification"
    for cell in ("A1", "B1"):
        ws[cell].font = header_font
        ws[cell].alignment = center

    col_idx = 3  # C 欄開始放產品欄位
    col_letter_map = {}
    for code in product_cols:
        info = product_info_map.get(code, {})
        gas = str(info.get("氣體", "") or "")
        ptype = str(info.get("Type", "") or "")
        status = str(info.get("出貨狀態", "") or "")
        qty = info.get("氣櫃數量", "")
        col_letter = get_column_letter(col_idx)
        ws[f"{col_letter}1"] = code
        ws[f"{col_letter}2"] = gas
        ws[f"{col_letter}3"] = ptype
        ws[f"{col_letter}4"] = status
        ws[f"{col_letter}5"] = f"Qty:{qty}"
        for r in range(1, 6):
            cell = ws[f"{col_letter}{r}"]
            cell.font = header_font
            cell.alignment = center
        ws.column_dimensions[col_letter].width = 12
        col_letter_map[code] = col_letter
        col_idx += 1

    summary_letters = {}
    for c in summary_cols:
        col_letter = get_column_letter(col_idx)
        ws.merge_cells(f"{col_letter}1:{col_letter}5")
        ws[f"{col_letter}1"] = c
        ws[f"{col_letter}1"].font = header_font
        ws[f"{col_letter}1"].alignment = center
        ws.column_dimensions[col_letter].width = 14
        summary_letters[c] = col_letter
        col_idx += 1

    row = 6
    for _, r in matrix_view.iterrows():
        ws[f"A{row}"] = r.get("Description", "")
        ws[f"B{row}"] = r.get("Specification", "")
        for code, col_letter in col_letter_map.items():
            ws[f"{col_letter}{row}"] = r.get(code, 0)
        for c, col_letter in summary_letters.items():
            val = r.get(c, "")
            ws[f"{col_letter}{row}"] = str(val) if val != "" else ""
        row += 1

    ws.freeze_panes = "C6"  # 固定左邊 Description/Specification + 上方多列表頭
    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 20

    def write_flat_sheet(sheet_name, df):
        s = wb.create_sheet(sheet_name)
        for j, col in enumerate(df.columns, start=1):
            cell = s.cell(row=1, column=j, value=col)
            cell.font = header_font
        for i, row_data in enumerate(df.itertuples(index=False), start=2):
            for j, val in enumerate(row_data, start=1):
                s.cell(row=i, column=j, value=str(val) if pd.notna(val) else "")
        for j in range(1, len(df.columns) + 1):
            s.column_dimensions[get_column_letter(j)].width = 16
        s.freeze_panes = "A2"

    write_flat_sheet("採購明細", purchase_df)
    write_flat_sheet("進貨明細", receipt_df)
    write_flat_sheet("產品規格總覽", products_info_df)

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def main():
    query_token = st.query_params.get("token", "")

    try:
        matrix, purchase, receipt, tokens, products_info = load_sheet_data()
    except Exception as e:
        st.error(f"資料讀取失敗，請稍後再試或聯絡窗口。({e})")
        st.stop()

    if not check_token(tokens, query_token):
        st.error("連結無效或已失效，請確認網址是否正確，或聯絡窗口取得最新連結。")
        st.stop()

    st.title("PA226009 專案 - 零件備貨進度")

    HIDDEN_COLS = ["Rev", "Item", "Remarks"]  # 零件品號保留在資料裡（不建欄位顯示），供點選列時識別用
    SUMMARY_COLS = ["總需求量", "累計採購數量", "累計進貨數量", "尚缺進貨量",
                     "最新採購日", "最新採購數量", "最新進貨日", "最新進貨數量"]

    product_info_map = products_info.set_index("產品品號").to_dict(orient="index")

    search = st.text_input("搜尋描述 / 規格", "")
    view = matrix.copy()
    if search:
        mask = (
            view["Description"].astype(str).str.contains(search, case=False, na=False)
            | view["Specification"].astype(str).str.contains(search, case=False, na=False)
        )
        view = view[mask]

    st.caption(f"共 {len(view)} 筆零件（全專案 {len(matrix)} 筆）")
    display_view = view.drop(columns=[c for c in HIDDEN_COLS if c in view.columns])

    product_cols = [c for c in display_view.columns if c in product_info_map]

    excel_buf = build_excel(
        display_view, purchase, receipt, products_info,
        product_cols, product_info_map,
        [c for c in SUMMARY_COLS if c in display_view.columns],
    )
    st.download_button(
        label="📥 匯出 Excel",
        data=excel_buf,
        file_name="PA226009_零件備貨進度.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

    # 主表格：Description/Specification 固定在左邊，
    # 每個產品欄位用巢狀分組做出「氣體 / Type / 出貨狀態 / Q'ty / 品號」多層表頭，比照 Excel 呈現方式
    column_defs = [
        {"field": "Description", "pinned": "left", "width": 220},
        {"field": "Specification", "pinned": "left", "width": 160},
    ]

    for code in product_cols:
        info = product_info_map.get(code, {})
        gas = str(info.get("氣體", "") or "")
        ptype = str(info.get("Type", "") or "")
        status = str(info.get("出貨狀態", "") or "")
        qty = info.get("氣櫃數量", "")
        column_defs.append({
            "headerName": code,
            "children": [{
                "headerName": gas,
                "children": [{
                    "headerName": ptype,
                    "children": [{
                        "headerName": status if status else "未標記出貨",
                        "children": [{
                            "field": code,
                            "headerName": f"Qty:{qty}",
                            "type": "numericColumn",
                            "width": 95,
                        }],
                    }],
                }],
            }],
        })

    DATE_SUMMARY_COLS = {"最新採購日", "最新進貨日"}
    for c in SUMMARY_COLS:
        if c in display_view.columns:
            width = 95 if c in DATE_SUMMARY_COLS else 90
            column_defs.append({"field": c, "width": width, "pinned": "right"})

    grid_options = {
        "columnDefs": column_defs,
        "defaultColDef": {"resizable": True, "sortable": True, "filter": False, "suppressMovable": True},
        "rowSelection": "single",
        "suppressRowClickSelection": False,
    }

    grid_response = AgGrid(
        display_view,
        gridOptions=grid_options,
        height=600,
        theme="streamlit",
        fit_columns_on_grid_load=False,
        allow_unsafe_jscode=True,
        update_mode=GridUpdateMode.SELECTION_CHANGED,
    )

    st.divider()
    st.subheader("查看單一零件的採購/進貨明細")
    st.caption("點選上方表格中的一列，即可看到該零件的採購/進貨明細")

    selected_rows = grid_response.get("selected_rows")
    # 不同版本的 streamlit-aggrid，selected_rows 可能是 DataFrame 或 list[dict]
    if selected_rows is not None and len(selected_rows) > 0:
        if isinstance(selected_rows, pd.DataFrame):
            selected_part = selected_rows.iloc[0]["零件品號"]
            selected_desc = selected_rows.iloc[0]["Description"]
        else:
            selected_part = selected_rows[0]["零件品號"]
            selected_desc = selected_rows[0]["Description"]
    else:
        selected_part = None
        selected_desc = None

    if selected_part:
        st.markdown(f"**已選擇：{selected_desc}**")
        col1, col2 = st.columns(2)

        with col1:
            st.markdown("**採購明細**")
            p_detail = purchase[purchase["零件品號"] == selected_part]
            if p_detail.empty:
                st.info("目前沒有採購紀錄")
            else:
                st.dataframe(
                    p_detail[["採購單號", "訂購數量", "預定交貨日"]],
                    use_container_width=True, hide_index=True,
                )

        with col2:
            st.markdown("**進貨明細**")
            r_detail = receipt[receipt["零件品號"] == selected_part]
            if r_detail.empty:
                st.info("目前沒有進貨紀錄")
            else:
                st.dataframe(
                    r_detail[["採購單號", "進貨數量", "進貨日期"]],
                    use_container_width=True, hide_index=True,
                )
    else:
        st.info("尚未選擇零件")


if __name__ == "__main__":
    main()
